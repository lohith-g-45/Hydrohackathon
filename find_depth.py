#!/usr/bin/env python3
"""Check all cells for 'depth' keyword."""

import openpyxl
import re

wb = openpyxl.load_workbook('HYDRO HACKATHON DATA.xlsx')

print("Searching for 'DEPTH' keyword in all sheets:\n")

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f"Sheet: {sheet_name}")
    for row in ws.iter_rows(min_row=1, max_row=40):
        for cell in row:
            if cell.value:
                cell_str = str(cell.value).lower()
                if 'depth' in cell_str:
                    # Get adjacent cell values
                    next_cell_value = cell.offset(0, 1).value
                    print(f"  Row {cell.row}: {cell.value} (next cell: {next_cell_value})")
    print()
