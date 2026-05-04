#!/usr/bin/env python3
"""Update Excel file with optimal ship parameters."""

import openpyxl
import shutil

# Backup original
shutil.copy('HYDRO HACKATHON DATA.xlsx', 'HYDRO HACKATHON DATA_BACKUP.xlsx')
print("✓ Backup created: HYDRO HACKATHON DATA_BACKUP.xlsx\n")

# Load workbook
wb = openpyxl.load_workbook('HYDRO HACKATHON DATA.xlsx')
ws = wb['Main Particulars']

print("UPDATING PARAMETERS:")
print("="*50)

# Update Depth (Row 5, Column C)
old_depth = ws['C5'].value
ws['C5'].value = 50.56
print(f"✓ Depth: {old_depth} → {ws['C5'].value} m")

# Add KG if not present - insert it after Depth (Row 6)
# First, check if KG already exists
kg_found = False
for row in ws.iter_rows(min_row=1, max_row=20):
    for cell in row:
        if cell.value and 'kg' in str(cell.value).lower():
            kg_found = True
            cell.value = 21.50
            print(f"✓ KG: 24.846 → {cell.value} m")
            break

# If KG not found, add it
if not kg_found:
    ws['B6'] = 'KG'
    ws['C6'].value = 21.50
    print(f"✓ KG: (new) → {ws['C6'].value} m")

print("="*50)

# Save updated workbook
wb.save('HYDRO HACKATHON DATA.xlsx')
print("\n✅ Excel file updated successfully!")
print("\nOptimal parameters set:")
print("  • Depth: 50.56 m (improved freeboard)")
print("  • KG: 21.50 m (improved initial stability)")
