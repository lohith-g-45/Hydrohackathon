#!/usr/bin/env python3
"""Update ship parameters in Excel to optimal values."""

from pathlib import Path
import openpyxl
from openpyxl.utils import get_column_letter
import shutil

# Backup original
shutil.copy('HYDRO HACKATHON DATA.xlsx', 'HYDRO HACKATHON DATA_BACKUP.xlsx')

# Load workbook
wb = openpyxl.load_workbook('HYDRO HACKATHON DATA.xlsx')

# Find and update parameters in "Ship Parameters" or similar sheet
for sheet in wb.sheetnames:
    ws = wb[sheet]
    print(f"\nScanning sheet: {sheet}")
    
    # Search for KG and Depth in the sheet
    for row in ws.iter_rows(min_row=1, max_row=50):
        for cell in row:
            if cell.value:
                cell_str = str(cell.value).lower()
                
                # Find and update KG
                if 'kg' in cell_str and ('center' in cell_str or cell.row < 15):
                    value_cell = cell.offset(0, 1)
                    if value_cell.value and isinstance(value_cell.value, (int, float)):
                        old_kg = value_cell.value
                        value_cell.value = 21.50
                        print(f"  ✓ Updated KG from {old_kg} to {value_cell.value}")
                
                # Find and update Depth
                if 'depth' in cell_str:
                    value_cell = cell.offset(0, 1)
                    if value_cell.value and isinstance(value_cell.value, (int, float)):
                        old_depth = value_cell.value
                        value_cell.value = 50.56
                        print(f"  ✓ Updated Depth from {old_depth} to {value_cell.value}")

# Save updated workbook
wb.save('HYDRO HACKATHON DATA.xlsx')
print("\n✅ Excel file updated with optimal parameters!")
print("   - KG: 21.50 m (improved stability)")
print("   - Depth: 50.56 m (improved deck immersion)")
