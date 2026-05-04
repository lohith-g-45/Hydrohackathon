#!/usr/bin/env python3
import openpyxl

wb = openpyxl.load_workbook('HYDRO HACKATHON DATA.xlsx')
ws = wb['Main Particulars']

print("Main Particulars Sheet:")
print("-" * 50)
for i in range(2, 10):
    b_val = ws[f'B{i}'].value
    c_val = ws[f'C{i}'].value
    if b_val or c_val:
        print(f"Row {i}: {b_val} = {c_val}")
